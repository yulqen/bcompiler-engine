import os
import zipfile
from pathlib import Path

import pytest
from lxml import etree
from openpyxl import load_workbook

from engine.parser.reader import SpreadsheetReader


@pytest.fixture
def heavy_template_datamap():
    here = os.path.abspath(os.curdir)
    return os.path.join(here, "tests/resources/org_templates/dft_datamap.csv")


@pytest.fixture
def template_containing_empty_sheet():
    here = os.path.abspath(os.curdir)
    return os.path.join(
        here, "tests/resources/test_template_with_introduction_sheet.xlsm"
    )


@pytest.fixture
def xml_test_file() -> Path:
    return Path.cwd() / "tests" / "resources" / "test.xml"


@pytest.fixture
def reader(org_test_files_dir, heavy_template_datamap) -> SpreadsheetReader:
    tmpl_file = org_test_files_dir / "dft1_tmp.xlsm"
    return SpreadsheetReader(tmpl_file, heavy_template_datamap)


def test_basic_xml_read(xml_test_file):
    tree = etree.parse(str(xml_test_file))
    assert (
        tree.getroot().tag
        == "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}workbook"
    )


def test_excel_reader_class(reader):
    assert isinstance(reader.archive, zipfile.ZipFile)


def test_excel_reader_class_file_list(reader):
    assert "xl/workbook.xml" in reader.valid_files


def test_excel_reader_class_has_package(reader):
    assert "/xl/worksheets/sheet22.xml" in reader.worksheet_files


def test_excel_reader_class_can_get_sheet_names(reader):
    assert reader.sheet_names[0] == "Introduction"


def test_excel_reader_class_can_get_shared_strings(reader):
    assert reader.shared_strings[0] == "Fantastic Portfolio Collection Sheet"


def test_excel_reader_class_can_get_rel_for_worksheet(reader):
    assert reader._get_sheet_rId("Introduction") == "rId3"
    assert reader._get_sheet_rId("Contents") == "rId4"


def test_excel_reader_class_can_get_worksheet_path_from_rId(reader):
    assert reader._get_worksheet_path_from_rId("rId3") == "worksheets/sheet1.xml"


def test_get_cell_value_for_cellref_sheet_lxml(reader):
    assert (
        reader.get_cell_value(cellref="C10", sheetname="Introduction")
        == "Coal Tits Ltd"
    )
    assert (
        reader.get_cell_value(cellref="C9", sheetname="Introduction")
        == "Institute of Hairdressing Dophins"
    )


def test_get_cell_values_for_sheet(reader):
    intro_vals = reader.get_cell_values("Introduction")
    assert intro_vals["B2"] == "Fantastic Portfolio Collection Sheet"
    assert intro_vals["B14"] == "Project Type (for GOASS use)"
    assert intro_vals["A35"] == "0.4.5"
    scope_vals = reader.get_cell_values("3 - Scope History")
    assert scope_vals["M41"] == "4th Scope Change"


def test_get_all_cell_vals_in_workbook(reader):
    sheets = reader.sheet_names
    vals = [reader.get_cell_values(sheetname) for sheetname in sheets]
    assert vals[0]["sheetname"] == "Introduction"
    assert vals[1]["sheetname"] == "Contents"
    assert vals[2]["sheetname"] == "Report Summary"
    assert vals[3]["sheetname"] == "GDPR - New SRO PDs"  # hidden sheet
    assert vals[4]["sheetname"] == "Data Quality Log"  # hidden sheet
    assert vals[5]["sheetname"] == "Dropdown List"  # hidden sheet
    assert vals[6]["sheetname"] == "Data Triangulation"  # hidden sheet
    assert vals[15]["sheetname"] == "4 - Leaders"  # hidden sheet


def test_raise_exception_when_trying_to_get_value_from_nonexistant_sheet(reader):
    with pytest.raises(ValueError):
        reader.get_cell_value(cellref="C1000", sheetname="Balls")


def test_return_none_when_cellref_out_of_range(reader):
    assert reader.get_cell_value(cellref="C1000", sheetname="Introduction") is None


############################################################
# THIS IS WHERE WE DO AN END-TO-END TEST FOR MASTER PARSING#
############################################################


def test_get_datamap_data_using_new_use_case(reader):
    """First test for gathering cell values in template using a datamap.

    Uses fast xpath parsing in lxml rather in openpyxl to extract the data.
    """
    template_data = reader.read()

    target_vals = [x.value for x in template_data[reader.fn]["Introduction"]]
    target_cellrefs = [x.cellref for x in template_data[reader.fn]["Introduction"]]
    assert "Universal Cantilever Bridge over the River Styx" in target_vals
    assert "C11" in target_cellrefs


def test_get_cell_values_using_lxml_handles_empty_sheet(
    template_containing_empty_sheet,
):
    reader = SpreadsheetReader(Path(template_containing_empty_sheet))
    data = reader.read_without_datamap()
    assert (
        len(data[reader.fn.parts[-1]]["data"]) == 2
    )  # we should not get an Introduction sheet as its empty


@pytest.mark.skip("used for exploring openpyxl")
def test_straight_read_using_openpyxl(org_test_files_dir):
    """The test template file here is full-sized and full of formatting.

    load_workbook() is very slow. Run with --profile
    """
    tmpl_file = org_test_files_dir / "dft1_tmp.xlsm"
    data = load_workbook(tmpl_file)
