"""
Test the processes involved in populating a bunch of blank
templates with data from a single master file.
"""

import datetime
from typing import List

import pytest
from engine.adapters.cli import write_master_to_templates
from engine.exceptions import MissingSheetFieldError
from engine.repository.templates import MultipleTemplatesWriteRepo
from engine.use_cases.output import WriteMasterToTemplates
from engine.utils.extraction import (ValidationReportItem,
                                     data_validation_report)
from openpyxl import load_workbook

# from openpyxl.worksheet.datavalidation import DataValidation


@pytest.mark.skip("We know this fails")
def test_write_into_dropdown(blank_org_template):
    wb = load_workbook(blank_org_template, read_only=False, keep_vba=True)
    ws = wb["1 - Project Info"]
    ws["E10"] = 101
    wb.save("/tmp/tosser.xlsm")
    wb1 = load_workbook("/tmp/tosser.xlsm", keep_vba=True)
    ws1 = wb1["1 - Project Info"]
    report: List[ValidationReportItem] = data_validation_report(ws1)
    formulae = [r.formula for r in report]
    assert '"Yes,No"' in formulae
    assert "'Drop Downs'!$F$3:$F$22" in formulae
    assert "'Drop Downs'!$S$3:$S$8" in formulae
    assert report[0].cell_range == "B6"
    assert report[2].cell_range.ranges[0].bounds == (2, 8, 2, 8
                                                     )  # represents B8
    assert report[2].cell_range.ranges[1].bounds == (2, 11, 2, 11
                                                     )  # represents B11


# NOT INCLUDED IN TESTS - FOR PROVING DATA VALIDATION
# def test_write_simple_data_validation_into_file():
#    wb = Workbook()
#    ws = wb.active
#    dv = DataValidation(type="list", formula1='"Trumpet,Piano"', allow_blank=True)
#    ws.add_data_validation(dv)
#    dv.add("A1:B20")
#    wb.save("/tmp/tosser.xlsx")
#
#

# def test_get_all_data_validation_in_sheet(blank_org_template):
#    wb = load_workbook(blank_org_template)
#    ws = wb["1 - Project Info"]
#    validations = ws.data_validations.dataValidation


@pytest.mark.skip("Are we doing the validation report?")
def test_validation_report(blank_org_template):
    wb = load_workbook(blank_org_template)
    ws = wb["1 - Project Info"]
    report = data_validation_report(ws)
    expected1 = 'Sheet: 1 - Project Info; E58:K58; Type: list; Formula: "Yes,No"'
    assert expected1 in report


def test_config_has_correct_files(mock_config):
    mock_config.initialise()
    input_dir = mock_config.PLATFORM_DOCS_DIR / "input"

    blank_fn = mock_config.config_parser["DEFAULT"]["blank file name"]
    datamap_fn = mock_config.config_parser["DEFAULT"]["datamap file name"]

    blank_template = input_dir / blank_fn
    datamap = input_dir / datamap_fn

    assert "tmp" in blank_template.parts
    assert "input" in blank_template.parts
    assert "Documents" in blank_template.parts
    assert "blank_template.xlsm" in blank_template.parts

    assert "tmp" in datamap.parts
    assert "input" in datamap.parts
    assert "Documents" in datamap.parts
    assert "datamap.csv" in datamap.parts


def test_exception_when_given_master_with_empty_col_a(mock_config, datamap,
                                                      master_no_col_a,
                                                      blank_template):
    """Test for handling cells in Col A which are empty - and return None."""
    mock_config.initialise()
    output_repo = MultipleTemplatesWriteRepo(blank_template)
    uc = WriteMasterToTemplates(output_repo, datamap, master_no_col_a,
                                blank_template)
    with pytest.raises(RuntimeError):
        uc.execute()


def test_can_export_more_than_twenty_six_columns(
    mock_config,
    datamap,
    master_with_rogue_cell_vals_beyond_col_and_row_range,
    blank_template,
):
    # master_plus_twenty_six must contain at least 26 projects
    mock_config.initialise()
    output_repo = MultipleTemplatesWriteRepo(blank_template)
    uc = WriteMasterToTemplates(
        output_repo,
        datamap,
        master_with_rogue_cell_vals_beyond_col_and_row_range,
        blank_template,
    )
    uc.execute()
    # Test a project much later - i.e. beyond the 25th project
    result_file = (mock_config.PLATFORM_DOCS_DIR / "output" /
                   "Ramsbottom Knot Gorge Cleanout 26.xlsm")
    wb = load_workbook(result_file)
    intro_sheet = wb["Introduction"]
    assert intro_sheet["C9"].value == "VA Department"


def test_output_gateway(mock_config, datamap, master, blank_template):
    mock_config.initialise()
    output_repo = MultipleTemplatesWriteRepo(blank_template)
    uc = WriteMasterToTemplates(output_repo, datamap, master, blank_template)
    uc.execute()
    result_file = mock_config.PLATFORM_DOCS_DIR / "output" / "Chutney Bridge.xlsm"
    wb = load_workbook(result_file)
    intro_sheet = wb["Introduction"]
    summ_sheet = wb["Summary"]
    an_sheet = wb["Another Sheet"]
    assert intro_sheet["C9"].value == "Accounting Department"
    assert intro_sheet["C10"].value == "Satellite Corp"
    assert intro_sheet["C13"].value == 1334
    # note that dates come out of spreadsheets as datetime objects
    assert intro_sheet["C17"].value == datetime.datetime(2012, 1, 1, 0, 0)
    assert summ_sheet["B3"].value == "Bobbins"
    assert an_sheet["F17"].value == 20.303
    # note that dates come out of spreadsheets as datetime objects
    assert an_sheet["H39"].value == datetime.datetime(2024, 2, 1, 0, 0)


def test_output_gateway_reduced_datamap(mock_config, datamap_reduced, master,
                                        blank_template):
    """
    The idea here is that the resulting master should work even if some lines
    in the  datamap are removed or hidden. In other words, the export does not
    depend on the datamap and the master's column A matching.
    """
    mock_config.initialise()
    output_repo = MultipleTemplatesWriteRepo(blank_template)
    uc = WriteMasterToTemplates(output_repo, datamap_reduced, master,
                                blank_template)
    uc.execute()
    result_file = mock_config.PLATFORM_DOCS_DIR / "output" / "Chutney Bridge.xlsm"
    wb = load_workbook(result_file)
    intro_sheet = wb["Introduction"]
    assert intro_sheet["C10"].value == "Satellite Corp"
    assert intro_sheet["C11"].value == "Chutney Bridge Ltd"
    assert intro_sheet["C13"].value == 1334
    # note that dates come out of spreadsheets as datetime objects
    assert intro_sheet["C17"].value == datetime.datetime(2012, 1, 1, 0, 0)


def test_export_continues_with_missing_sheet_in_datamap(
        mock_config, master, datamap_missing_fields, blank_template):
    """When an export takes place using a datamap with missing sheet names"""
    mock_config.initialise()
    output_repo = MultipleTemplatesWriteRepo(blank_template)
    uc = WriteMasterToTemplates(output_repo, datamap_missing_fields, master,
                                blank_template)
    with pytest.raises(MissingSheetFieldError):
        uc.execute()
