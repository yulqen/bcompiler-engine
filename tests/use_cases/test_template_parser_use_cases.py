import json
import shutil
from pathlib import Path

import pytest
from engine.exceptions import NestedZipError
from engine.repository.datamap import InMemorySingleDatamapRepository
from engine.repository.master import MasterOutputRepository
from engine.repository.templates import (
    FSPopulatedTemplatesRepo,
    InMemoryPopulatedTemplatesRepository,
    InMemoryPopulatedTemplatesZip,
)
from engine.use_cases.parsing import (
    ApplyDatamapToExtractionUseCase,
    CreateMasterUseCase,
    ParsePopulatedTemplatesUseCase,
)
from engine.utils.extraction import _check_file_in_datafile
from openpyxl import load_workbook


def test_template_parser_use_case(resources):
    repo = InMemoryPopulatedTemplatesRepository(resources)
    parse_populated_templates_use_case = ParsePopulatedTemplatesUseCase(repo)
    result = parse_populated_templates_use_case.execute()
    assert (
        json.loads(result)["test_template.xlsx"]["data"]["Summary"]["B3"]["value"]
        == "This is a string"
    )


def test_query_data_from_data_file(
    mock_config, dat_file, spreadsheet_same_data_as_dat_file
):
    mock_config.initialise()
    shutil.copy2(dat_file, mock_config.DATAMAPS_LIBRARY_DATA_DIR)
    shutil.copy2(
        spreadsheet_same_data_as_dat_file, mock_config.PLATFORM_DOCS_DIR / "input"
    )
    repo = FSPopulatedTemplatesRepo(mock_config.PLATFORM_DOCS_DIR)
    parse_populated_templates_use_case = ParsePopulatedTemplatesUseCase(repo)
    result = parse_populated_templates_use_case.execute()
    assert (
        json.loads(result)["test_dat_file_use_case.xlsx"]["data"]["new sheet"]["A1"][
            "value"
        ]
        == "Project/Programme Name"
    )


def test_zip_with_directory_raises_exception(
    mock_config, datamap, templates_zipped_containing_dir
):
    mock_config.initialise()
    shutil.copy2(datamap, (Path(mock_config.PLATFORM_DOCS_DIR) / "input"))
    zip_repo = InMemoryPopulatedTemplatesZip(templates_zipped_containing_dir)
    dm_repo = InMemorySingleDatamapRepository(
        Path(mock_config.PLATFORM_DOCS_DIR) / "input" / "datamap.csv"
    )
    uc = ApplyDatamapToExtractionUseCase(dm_repo, zip_repo)
    with pytest.raises(NestedZipError):
        uc.execute()


def test_extract_data_from_templates_in_zip_file(
    mock_config, datamap_match_test_template, templates_zipped
):
    mock_config.initialise()
    shutil.copy2(
        datamap_match_test_template, (Path(mock_config.PLATFORM_DOCS_DIR) / "input")
    )
    zip_repo = InMemoryPopulatedTemplatesZip(templates_zipped)
    dm_repo = InMemorySingleDatamapRepository(
        Path(mock_config.PLATFORM_DOCS_DIR)
        / "input"
        / "datamap_match_test_template.csv"
    )
    uc = ApplyDatamapToExtractionUseCase(dm_repo, zip_repo)
    uc.execute()
    assert (
        uc.query_key(
            "test_template_with_introduction_sheet.xlsm", "String Key", "Summary"
        )
        == "This is a string"
    )
    assert (
        uc.query_key(
            "test_template_with_introduction_sheet.xlsm", "Big Float", "Another Sheet"
        )
        == 7.2
    )
    assert "test_template2.xlsx" in uc._template_data_dict.keys()
    assert uc.query_key("test_template2.xlsx", "Big Float", "Another Sheet") == 7.2


@pytest.mark.slow
def test_in_memory_datamap_application_to_extracted_data(
    mock_config, datamap, template_with_introduction_sheet
):
    mock_config.initialise()
    shutil.copy2(
        template_with_introduction_sheet,
        (Path(mock_config.PLATFORM_DOCS_DIR) / "input"),
    )
    shutil.copy2(datamap, (Path(mock_config.PLATFORM_DOCS_DIR) / "input"))
    tmpl_repo = InMemoryPopulatedTemplatesRepository(
        mock_config.PLATFORM_DOCS_DIR / "input"
    )
    dm_repo = InMemorySingleDatamapRepository(
        Path(mock_config.PLATFORM_DOCS_DIR) / "input" / "datamap.csv"
    )
    uc = ApplyDatamapToExtractionUseCase(dm_repo, tmpl_repo)
    uc.execute()
    assert (
        uc.query_key(
            "test_template_with_introduction_sheet.xlsm", "String Key", "Summary"
        )
        == "This is a string"
    )
    assert (
        uc.query_key(
            "test_template_with_introduction_sheet.xlsm", "Big Float", "Another Sheet"
        )
        == 7.2
    )


def test_in_memory_datamap_application_to_extracted_data_raises_exception(
    mock_config, datamap, template
):
    "Raise exception when the key provided is not in the datamap"
    mock_config.initialise()
    shutil.copy2(template, (Path(mock_config.PLATFORM_DOCS_DIR) / "input"))
    tmpl_repo = InMemoryPopulatedTemplatesRepository(
        mock_config.PLATFORM_DOCS_DIR / "input"
    )
    dm_repo = InMemorySingleDatamapRepository(datamap)
    uc = ApplyDatamapToExtractionUseCase(dm_repo, tmpl_repo)
    with pytest.raises(KeyError):
        # note the extra space in the key name
        uc.query_key("test_template.xlsx", "Funny Date ", "Another Sheet")
    with pytest.raises(KeyError):
        # note the extra space in the sheet name
        uc.query_key("test_template.xlsx", "Funny Date", "Another Sheet ")


def test_in_memory_datamap_generator(
    mock_config, datamap_match_test_template, template
):
    "Doesn't really need a generator because its already in memory, but whatever..."
    mock_config.initialise()
    shutil.copy2(template, (Path(mock_config.PLATFORM_DOCS_DIR) / "input"))
    tmpl_repo = InMemoryPopulatedTemplatesRepository(
        mock_config.PLATFORM_DOCS_DIR / "input"
    )
    dm_repo = InMemorySingleDatamapRepository(datamap_match_test_template)
    uc = ApplyDatamapToExtractionUseCase(dm_repo, tmpl_repo)
    uc.execute()
    data = uc.get_values()
    #   assert next(uc.get_values(as_obj=True)) == {("test_template.xlsx", "Summary", "B2"): datetime.date(2019, 10, 19)}
    assert next(data) == {
        ("test_template.xlsx", "Date Key", "Summary", "B2"): "2019-10-20T00:00:00"
    }
    assert next(data) == {
        ("test_template.xlsx", "String Key", "Summary", "B3"): "This is a string"
    }
    assert next(data) == {
        ("test_template.xlsx", "Big Float", "Another Sheet", "F17"): 7.2
    }


def test_create_master_spreadsheet(mock_config, datamap_match_test_template, template):
    mock_config.initialise()
    shutil.copy2(template, (Path(mock_config.PLATFORM_DOCS_DIR) / "input"))
    tmpl_repo = InMemoryPopulatedTemplatesRepository(
        mock_config.PLATFORM_DOCS_DIR / "input"
    )
    dm_repo = InMemorySingleDatamapRepository(datamap_match_test_template)
    output_repo = MasterOutputRepository
    uc = CreateMasterUseCase(dm_repo, tmpl_repo, output_repo)
    uc.execute("master.xlsx")
    wb = load_workbook(Path(mock_config.PLATFORM_DOCS_DIR) / "output" / "master.xlsx")
    ws = wb.active
    assert ws["A1"].value == "file name"
    assert ws["B1"].value == "test_template"
    assert ws["B2"].value == "2019-10-20T00:00:00"
    assert ws["B3"].value == "This is a string"


def ensure_data_and_populate_file(config, dat_file, spreadsheet_file) -> None:
    "Ensure the data in a single file is mirrored in a dat file, in correct location for testing"
    shutil.copy2(dat_file, config.DATAMAPS_LIBRARY_DATA_DIR)
    shutil.copy2(spreadsheet_file, config.PLATFORM_DOCS_DIR)


def test_file_data_is_in_dat_file(
    mock_config, dat_file, spreadsheet_same_data_as_dat_file
):
    """Before we do any data extraction from files, we need to check out dat file."""
    mock_config.initialise()
    ensure_data_and_populate_file(
        mock_config, dat_file, spreadsheet_same_data_as_dat_file
    )
    assert _check_file_in_datafile(spreadsheet_same_data_as_dat_file, dat_file)


def test_file_data_is_in_dat_file_works_with_str_params(
    mock_config, dat_file, spreadsheet_same_data_as_dat_file
):
    """Before we do any data extraction from files, we need to check out dat file."""
    mock_config.initialise()
    ensure_data_and_populate_file(
        mock_config, dat_file, spreadsheet_same_data_as_dat_file
    )
    assert _check_file_in_datafile(
        str(spreadsheet_same_data_as_dat_file), str(dat_file)
    )


def test_if_data_file_and_spreadsheet_file_dont_exist():
    with pytest.raises(FileNotFoundError) as exinfo:
        _check_file_in_datafile("no_spreadsheet_file.xlsx", "no_dat_file.xlsx")
    msg = exinfo.value.args[0]
    assert "Cannot find" in msg


def test_file_data_is_in_data_file__but_spreadsheet_file_doesnt_exist(
    mock_config, dat_file
):
    mock_config.initialise()
    with pytest.raises(FileNotFoundError):
        _check_file_in_datafile("/home/non-existant-file.txt", dat_file)


def test_file_data_not_in_data_returns_exception(
    mock_config, dat_file, spreadsheet_one_cell_different_data_than_dat_file
):
    """Before we do any data extraction from files, we need to check out dat file.

    If file data is not contained in the dat file, we need to know about it by getting a False return.
    """
    mock_config.initialise()
    ensure_data_and_populate_file(
        mock_config, dat_file, spreadsheet_one_cell_different_data_than_dat_file
    )
    with pytest.raises(KeyError) as excinfo:
        _check_file_in_datafile(
            spreadsheet_one_cell_different_data_than_dat_file, dat_file
        )
    exception_msg = excinfo.value.args[0]
    assert (
        exception_msg
        == "Data from test_data_file_use_case_diff_data_from_dat_file.xlsx is not contained in extracted_data.dat"
    )


def test_in_extract_files_from_zipfile(mock_config, datamap, templates_zipped):
    mock_config.initialise()
    shutil.copy2(datamap, (Path(mock_config.PLATFORM_DOCS_DIR) / "input"))
    tmpl_repo = InMemoryPopulatedTemplatesZip(templates_zipped)
    dm_repo = InMemorySingleDatamapRepository(
        Path(mock_config.PLATFORM_DOCS_DIR) / "input" / "datamap.csv"
    )
    uc = ApplyDatamapToExtractionUseCase(dm_repo, tmpl_repo)
    uc.execute()
    assert (
        uc.query_key(
            "test_template_with_introduction_sheet.xlsm", "String Key", "Summary"
        )
        == "This is a string"
    )
    assert (
        uc.query_key(
            "test_template_with_introduction_sheet.xlsm", "Big Float", "Another Sheet"
        )
        == 7.2
    )
    assert (
        uc.query_key(
            "test_template_with_introduction_sheet.XLSM", "Big Float", "Another Sheet"
        )
        == 7.2
    )
    assert (
        uc.query_key(
            "test_template_with_introduction_sheet.XLSX", "Big Float", "Another Sheet"
        )
        == 7.2
    )


def test_in_extract_files_from_zipfile_with_deep_structure_raises_exception(
    mock_config, datamap, templates_zipped_deep_structure
):
    mock_config.initialise()
    shutil.copy2(datamap, (Path(mock_config.PLATFORM_DOCS_DIR) / "input"))
    tmpl_repo = InMemoryPopulatedTemplatesZip(templates_zipped_deep_structure)
    dm_repo = InMemorySingleDatamapRepository(
        Path(mock_config.PLATFORM_DOCS_DIR) / "input" / "datamap.csv"
    )
    uc = ApplyDatamapToExtractionUseCase(dm_repo, tmpl_repo)
    with pytest.raises(NestedZipError):
        uc.execute()
