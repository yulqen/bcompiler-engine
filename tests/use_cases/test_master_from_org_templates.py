# WARNING: slow tests!
import os
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from engine.repository.datamap import InMemorySingleDatamapRepository
from engine.repository.master import MasterOutputRepository
from engine.repository.templates import InMemoryPopulatedTemplatesRepository
from engine.use_cases.parsing import CreateMasterUseCase
from engine.utils.extraction import datamap_reader, template_reader


@pytest.mark.slow
def test_datamap_reader(mock_config, org_test_files_dir):
    mock_config.initialise()
    for fl in os.listdir(org_test_files_dir):
        shutil.copy(
            Path.cwd() / "tests" / "resources" / "org_templates" / fl,
            (Path(mock_config.PLATFORM_DOCS_DIR) / "input"),
        )
    dm_file = mock_config.PLATFORM_DOCS_DIR / "input" / "dft_datamap.csv"
    data = datamap_reader(dm_file)
    assert data[0].key == "Project/Programme Name"
    assert data[0].sheet == "Introduction"
    assert data[0].cellref == "C11"
    assert data[0].filename == str(dm_file)


@pytest.mark.slow
def test_template_reader(mock_config, org_test_files_dir):
    mock_config.initialise()
    for fl in os.listdir(org_test_files_dir):
        shutil.copy(
            Path.cwd() / "tests" / "resources" / "org_templates" / fl,
            (Path(mock_config.PLATFORM_DOCS_DIR) / "input"),
        )
    tmpl_file = Path(mock_config.PLATFORM_DOCS_DIR) / "input" / "dft1_tmp.xlsm"
    data = template_reader(tmpl_file)
    assert data["dft1_tmp.xlsm"]["data"]["10 - Benefits"]["X34"]["value"] == 1
    assert (
        data["dft1_tmp.xlsm"]["data"]["10 - Benefits"]["X34"]["data_type"] == "NUMBER"
    )


@pytest.mark.slow
def test_create_master_spreadsheet(mock_config, org_test_files_dir):
    mock_config.initialise()
    for fl in os.listdir(org_test_files_dir):
        shutil.copy(
            Path.cwd() / "tests" / "resources" / "org_templates" / fl,
            (Path(mock_config.PLATFORM_DOCS_DIR) / "input"),
        )
    tmpl_repo = InMemoryPopulatedTemplatesRepository(
        mock_config.PLATFORM_DOCS_DIR / "input"
    )
    dm_file = mock_config.PLATFORM_DOCS_DIR / "input" / "dft_datamap.csv"
    dm_repo = InMemorySingleDatamapRepository(str(dm_file))
    output_repo = MasterOutputRepository
    uc = CreateMasterUseCase(dm_repo, tmpl_repo, output_repo)
    uc.execute("master.xlsx")
    wb = load_workbook(Path(mock_config.PLATFORM_DOCS_DIR) / "output" / "master.xlsx")
    ws = wb.active
    assert ws["A1"].value == "file name"
    assert "dft1_tmp" in ws["B1"].value
    # reintroduce this test once can be locked


#   assert ws["B2"].value == "2019-10-20T00:00:00"
#   assert ws["B3"].value == "This is a string"
