from pathlib import Path

from openpyxl import load_workbook

from engine.repository.master import MasterOutputRepository


def test_master_output_repository(mock_config):
    mock_config.initialise()
    test_data = [
        {("test_file.xlsx", "Test Key", "Test Sheet", "A1"): "Test Value"}
    ]
    repo = MasterOutputRepository(test_data, "test_master.xlsx")
    repo.save()
    wb = load_workbook((Path(mock_config.PLATFORM_DOCS_DIR) / "output" / "test_master.xlsx"))
    ws = wb.active
    assert ws.title == "Master"
    assert ws["A1"].value == "file name"
    assert ws["B1"].value == "test_file"
    assert ws["A2"].value == "Test Key"
    assert ws["B2"].value == "Test Value"
