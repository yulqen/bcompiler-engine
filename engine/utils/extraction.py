import csv
import datetime
import enum
import fnmatch
import hashlib
import json
import logging
import os
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from itertools import groupby
from pathlib import Path
from typing import Any, Dict, List, NamedTuple, Union
from zipfile import BadZipFile

from engine.config import Config  # noqa
from engine.domain.datamap import DatamapFile, DatamapLine, DatamapLineValueType
from engine.domain.template import TemplateCell
from engine.exceptions import (
    DatamapFileEncodingError,
    DatamapNotCSVException,
    MalFormedCSVHeaderException,
    MissingCellKeyError,
    MissingLineError,
    MissingSheetFieldError,
    NoApplicableSheetsInTemplateFiles,
)
from engine.utils import ECHO_FUNC_GREEN, ECHO_FUNC_YELLOW
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.worksheet import Worksheet

FILE_DATA = Dict[str, Union[str, Dict[str, Dict[str, str]]]]
DAT_DATA = Dict[str, FILE_DATA]
SHEET_DATA_IN_LST = List[Dict[str, str]]
ALL_IMPORT_DATA = Dict[str, Dict[str, Dict[str, Dict[str, Dict[str, str]]]]]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s: %(levelname)s - %(message)s",
    datefmt="%d-%b-%y %H:%M:%S",
)
logger = logging.getLogger(__name__)


def _dml_line_check(line: OrderedDict, headers: Dict[str, str]) -> None:
    """Checks a line from csv.DictReader and raises exception if isn't complete.

    Otherwise passes.
    """
    # Missing sheet field check
    # if we have a blank sheet field
    missing_fields = [x[0] for x in line.items() if x[1] == ""]
    if (
        headers.get("key") in missing_fields
        and headers.get("sheet") in missing_fields
        and headers.get("cellref") in missing_fields
    ):
        raise MissingLineError(
            "Datamap contains a missing line. Please fix datamap before proceeding."
        )
    if headers.get("sheet") in missing_fields:
        raise MissingSheetFieldError(
            f"Line whose key is {line['cell_key']} is missing a sheet field. Cannot proceed."
            f" Please fix datamap."
        )
    if headers.get("key") in missing_fields:
        raise MissingCellKeyError(
            f"Line missing a key field. Contains sheet {line['template_sheet']} and cell reference {line['cellreference']}. "
            f"Cannot proceed."
            f" Please fix datamap."
        )


def datamap_reader(dm_file: Union[Path, str]) -> List[DatamapLine]:
    "Given a datamap csv file, returns a list of DatamapLine objects."
    try:
        headers = datamap_check(dm_file)
    except DatamapNotCSVException:
        raise
    data = []
    logger.info(f"Reading datamap {dm_file}")
    logger.info("Checking that datamap is valid.")
    with DatamapFile(dm_file) as datamap_file:
        reader = csv.DictReader(datamap_file)
        for line in reader:
            # if dml is correct, this passes silently
            _dml_line_check(line, headers)
            try:
                if headers["type"] is None:
                    data.append(
                        DatamapLine(
                            key=_clean(line[headers["key"]]),
                            sheet=_clean(line[headers["sheet"]]),
                            cellref=_clean(line[headers["cellref"]], is_cellref=True),
                            data_type=None,
                            filename=str(dm_file),
                        )
                    )
                else:
                    data.append(
                        DatamapLine(
                            key=_clean(line[headers["key"]]),
                            sheet=_clean(line[headers["sheet"]]),
                            cellref=_clean(line[headers["cellref"]], is_cellref=True),
                            data_type=_clean(line[headers["type"]]),
                            filename=str(dm_file),
                        )
                    )
            except AttributeError as e:
                if "Cannot clean value other than a string here." in e.args:
                    logger.warning(
                        f"{line[headers['key']]} line in datamap may be missing a key field, or"
                        f" the line is formed unexpectedly. This line will be skipped during import/export."
                        f" Check your datamap!"
                    )
    return data


class CheckType(enum.Enum):
    FAIL = enum.auto()
    PASS = enum.auto()
    MISSING_SHEETS_REQUIRED_BY_DATAMAP = enum.auto()
    UNDEFINED = enum.auto()


@dataclass
class Check:
    filename: str
    sheet: str
    proceed: bool = False
    state: CheckType = CheckType.UNDEFINED
    error_type: CheckType = CheckType.UNDEFINED
    msg: str = ""


def remove_failing_files(
    lst_of_checks: List[Check], template_data: ALL_IMPORT_DATA
) -> ALL_IMPORT_DATA:
    """Given a list of checks, identify files which contain CheckType.FAIL, then remove them from template_data.

    If this results in an empty template_data dict, return None.
    """
    failing_files = list(
        set([(f.filename, f.sheet) for f in lst_of_checks if f.state == CheckType.FAIL])
    )
    for f in list(set([x[0] for x in failing_files])):
        missing_sheets = [x.sheet for x in lst_of_checks if x.filename == f]
        for ms in missing_sheets:
            logger.warning(
                f"{ms} sheet missing from {f} - it is required by the datamap."
            )
        template_data.pop(f)
        logger.warning(
            f"{f} skipped due to not having requisite sheets named in datamap."
        )
    #       raise RemoveFileWithNoSheetRequiredByDatamap(f)
    if len(template_data.keys()) > 1:
        return template_data
    elif len(template_data.keys()) < 1:
        msg = "There are no files containing sheets declared in datamap. Quitting."
        logger.critical(msg)
        raise NoApplicableSheetsInTemplateFiles(msg)
    return template_data


def check_datamap_sheets(
    datamap_data: List[Dict[str, str]], template_data: ALL_IMPORT_DATA
) -> List[Check]:
    "Parse data struct for each of datamap and all template data for sheet compliance."
    checks = []
    sheets_in_datamap: List[str] = list(set([x["sheet"] for x in datamap_data]))
    files_in_template_data = list(template_data.keys())
    sheets_in_template_data = {
        x: list(template_data[x]["data"].keys()) for x in files_in_template_data
    }
    for f in files_in_template_data:
        for s in sheets_in_datamap:
            if s in sheets_in_template_data[f]:
                checks.append(
                    Check(
                        filename=f,
                        sheet=s,
                        proceed=True,
                        state=CheckType.PASS,
                        error_type=CheckType.UNDEFINED,
                        msg=f"File {f} checked: OK.",
                    )
                )
            else:
                checks.append(
                    Check(
                        filename=f,
                        sheet=s,
                        proceed=False,
                        state=CheckType.FAIL,
                        error_type=CheckType.MISSING_SHEETS_REQUIRED_BY_DATAMAP,
                        msg=f"File {f} has no sheet[s] {s}.",
                    )
                )
    return checks


class ValidationReportItem(NamedTuple):
    """Can be used to allow better access to data from a Data Validation."""

    formula: str
    cell_range: MultiCellRange
    report_line: str


def data_validation_report(sheet: Worksheet) -> List[ValidationReportItem]:
    """Given an openpyxl sheet, produces a list of statements regarding dv cells.
    To be used by the adapters.
    :type sheet: Worksheet
    """
    output = []
    validations = sheet.data_validations.dataValidation
    for v in validations:
        output.append(
            ValidationReportItem(
                v.formula1,
                v.ranges,
                f"Sheet: {sheet.title}; {v.sqref}; Type: {v.type}; Formula: {v.formula1}",
            )
        )
    return output


def _check_file_in_datafile(spreadsheet_file: Path, data_file: Path) -> bool:
    """Given a spreadsheet file, checks whether its data is already contained in a data file.
    - Raises KeyError if file not found in data file.
    - Raises FileNotFoundError if either the spreadsheet_file or data_file files cannot be found.
    :type spreadsheet_file: Path
    :type data_file: Path
    """
    # expect the params to be Paths, so we convert
    spreadsheet_file = Path(spreadsheet_file)
    data_file = Path(data_file)
    if not spreadsheet_file.is_file():
        raise FileNotFoundError("Cannot find {}".format(str(spreadsheet_file)))
    if not data_file.is_file():
        raise FileNotFoundError("Cannot find {}".format(str(data_file)))
    f_checksum = _hash_single_file(spreadsheet_file)
    with open(data_file, encoding="utf-8") as f:
        data: DAT_DATA = json.load(f)
        try:
            f_data: FILE_DATA = data[spreadsheet_file.name]
            return f_data["checksum"] == f_checksum
        except KeyError:
            raise KeyError(
                "Data from {} is not contained in {}".format(
                    spreadsheet_file.name, data_file.name
                )
            )


def get_xlsx_files(directory: Path) -> List[Path]:
    """Return a list of Path objects for each xlsx file in directory, or raise an exception."""
    output = []
    if not os.path.isabs(directory):
        raise RuntimeError("Require absolute path here")
    for file_path in os.listdir(directory):
        if (
            fnmatch.fnmatch(file_path, "*.xls[xm]")
            and "blank_template" not in file_path
        ):
            output.append(Path(os.path.join(directory, file_path)))
    return output


def _get_cell_data(filepath: Path, data, sheet_name: str, cellref: str):
    """Given a Path and a dict of data - and a sheet name, AND a cellref..

    Return:
        - a dict representing the TemplateCell in string format
    """
    _file_data = data[filepath.name]["data"]
    return _file_data[sheet_name][cellref]


def _clean(target_str: str, is_cellref: bool = False) -> str:
    """Rids a string of its most common problems: spacing, capitalisation,etc."""
    try:
        output_str = target_str.lstrip().rstrip()
    except AttributeError:
        raise AttributeError("Cannot clean value other than a string here.")
    if is_cellref:
        output_str = output_str.upper()
    return output_str


def _extract_sheets(lst_of_tcs: List[TemplateCell]) -> Dict[str, List[TemplateCell]]:
    """Given a list of TemplateCell objects, returns the list but as a dict sorted by its sheet_name"""
    output: Dict[str, List[TemplateCell]] = {}
    data = sorted(lst_of_tcs, key=lambda x: x.sheet_name)
    for k, group in groupby(data, key=lambda x: x.sheet_name):
        output.update({k: list(group)})
    return output


def _extract_cellrefs(lst_of_tcs: SHEET_DATA_IN_LST):
    """Extract value from TemplateCell.cellref for each TemplateCell in a list to group them.

    When given a list of TemplateCell objects, this function extracts each TemplateCell
    by it's cellref value and groups them according. In the current implementation, this is
    only called on a list of TemplateCell objects which have the same sheet_name value, and
    therefore expects to find only a single cellref value each time, meaning that the list
    produced by groupby() can be removed and the single value return. Returns an exception
    if this list has more than one object.

    Args:
        lst_of_tcs: List of TemplateCell objects.

    Raises:
        RuntimeError: if more than one cellref value is found in the list.

    Returns:
        Dictionary whose key is the cellref and value is the TemplateCell that contains it.

    """
    output = {}
    data = sorted(lst_of_tcs, key=lambda x: x["cellref"])
    for k, group in groupby(data, key=lambda x: x["cellref"]):
        result_lst = list(group)
        if len(result_lst) > 1:
            raise RuntimeError(
                "Found duplicate sheet/cellref item when extracting keys."
            )
        result_item = result_lst[0]
        output.update({k: result_item})
    return output


def _hash_single_file(filepath: Path) -> str:
    """Return a checksum for a given file at Path.

    Returns checksum string.

    Raises FileNotFoundError if cannot find filepath.
    """
    # if we're given a str, we convert
    filepath = Path(filepath)
    try:
        filepath.is_file()
    except FileNotFoundError:
        raise FileNotFoundError(
            "Cannot find {} in order to calculate checksum".format(filepath)
        )
    hash_obj = hashlib.md5(open(filepath, "rb").read())
    return hash_obj.digest().hex()


def _hash_target_files(list_of_files: List[Path]) -> Dict[str, str]:
    """Hash each file in list_of_files.

    Given a list of files, return a dict containing the file name as
    keys and md5 hash as value for each file.
    """
    output = {}
    for file_name in list_of_files:
        if os.path.isfile(file_name):
            hash_obj = hashlib.md5(open(file_name, "rb").read())
            output.update({file_name.name: hash_obj.digest().hex()})
    return output


def datamap_check(dm_file):
    """Given a datamap csv file, returns a dict of the headers used in reality...

    raises IndexError if less than three headers are found (type header can be None)
    """
    if ECHO_FUNC_YELLOW is not None:
        ECHO_FUNC_YELLOW("Checking datamap file {}\n".format(dm_file))
    _good_keys = ["cell_key", "cellkey", "key"]
    _good_sheet = ["template_sheet", "sheet", "templatesheet"]
    _good_cellref = ["cell_reference", "cell_ref", "cellref", "cellreference"]
    _good_type = ["type", "value_type", "cell_type", "celltype"]
    headers = {}
    using_type = True
    try:
        with DatamapFile(dm_file) as datamap_file:
            # initial check - have we got enough headers? If not - raise exception
            top_row = next(datamap_file).rstrip().split(",")
            if len(top_row) == 1:
                # test for first char being ascii - if not, likely wrong encoding
                if not top_row[0][0].isascii():
                    raise DatamapFileEncodingError(
                        "Incorrect encoding of datamap file. Please ensure "
                        "it is saved in Excel using CSV (Comma delimited) type - not CSV UTF-8 (Comma delimited) type."
                    )
                else:
                    raise MalFormedCSVHeaderException(
                        "Datamap contains only one header - need at least three to proceed. Quitting."
                    )
            if len(top_row) == 2:
                raise MalFormedCSVHeaderException(
                    "Datamap contains only two headers - need at least three to proceed. Quitting."
                )
            if not top_row[0][0].isascii():
                raise DatamapFileEncodingError(
                    "Incorrect encoding of datamap file. Please ensure "
                    "it is saved in Excel using CSV (Comma delimited) type - not CSV UTF-8 (Comma delimited) type."
                )
            if top_row[-1] not in _good_type:
                # test if we are using type column here
                headers.update(type=None)
                using_type = False
            if top_row[0] in _good_keys:
                headers.update(key=top_row[0])
            if top_row[1] in _good_sheet:
                headers.update(sheet=top_row[1])
            if top_row[2] in _good_cellref:
                headers.update(cellref=top_row[2])
            if using_type:
                if top_row[3] in _good_type:
                    headers.update(type=top_row[3])
    except DatamapNotCSVException:
        raise
    if len(headers.keys()) >= 2:
        # final test - we don't want to proceed unless we have minimum headers
        if not all(
            [x in list(headers.keys()) for x in ["key", "sheet", "cellref", "type"]]
        ):
            raise MalFormedCSVHeaderException(
                "Cannot proceed without required number of headers"
            )
        if ECHO_FUNC_GREEN is not None:
            ECHO_FUNC_GREEN("{} checked ok\n".format(dm_file))
        return headers
    else:
        return MalFormedCSVHeaderException(
            "Datamap does not contain the required headers. Cannot proceed"
        )


def template_reader(template_file) -> Dict[str, Dict[str, Dict[Any, Any]]]:
    """Given a populated xlsx file, returns all data in a list of TemplateCell objects

    This test uses a fully formatted template file.
    ."""
    logger.info(f"Starting import of {template_file}.")
    inner_dict: Dict[str, Dict[Any, Any]] = {"data": {}}
    f_path: Path = Path(template_file)
    try:
        workbook = load_workbook(template_file, data_only=True)
    except TypeError:
        msg = (
            "Unable to open {}. Potential corruption of file. Try resaving "
            "in Excel or removing conditional formatting. See issue at "
            "https://github.com/hammerheadlemon/bcompiler-engine/issues/3 for update. Quitting.".format(
                f_path
            )
        )
        logger.critical(msg)
        raise
    except BadZipFile:
        logger.critical(
            f"Cannot open {template_file} due to file not conforming to expected format. "
            f"Not continuing. Remove file from input directory and try again."
        )
        raise RuntimeError
    checksum: str = _hash_single_file(f_path)
    holding = []
    for sheet in workbook.worksheets:
        sheet_data: SHEET_DATA_IN_LST = []
        sheet_dict: Dict[str, Dict[str, Dict[str, str]]] = {}
        rowcnt = 0
        for row in sheet.rows:
            if rowcnt > int(Config.TEMPLATE_ROW_LIMIT):
                break
            for cell in row:
                if cell.value is not None:
                    try:
                        val = cell.value.rstrip().lstrip()
                        c_type = DatamapLineValueType.TEXT
                    except AttributeError:
                        if isinstance(cell.value, (float, int)):
                            val = cell.value
                            c_type = DatamapLineValueType.NUMBER
                        elif isinstance(cell.value, (datetime.date, datetime.datetime)):
                            val = cell.value.isoformat()
                            c_type = DatamapLineValueType.DATE
                    cellref = "{}{}".format(cell.column_letter, cell.row)
                    if isinstance(template_file, Path):
                        t_cell = TemplateCell(
                            template_file.as_posix(), sheet.title, cellref, val, c_type
                        ).to_dict()
                    else:
                        t_cell = TemplateCell(
                            template_file, sheet.title, cellref, val, c_type
                        ).to_dict()
                    sheet_data.append(t_cell)
            rowcnt += 1
        sheet_dict.update({sheet.title: _extract_cellrefs(sheet_data)})
        holding.append(sheet_dict)
    for sd in holding:
        inner_dict["data"].update(sd)
        inner_dict.update({"checksum": checksum})  # type: ignore
    shell_dict = {f_path.name: inner_dict}
    logger.info(f"Compiled data from {f_path.name}")
    return shell_dict


def max_tmpl_row(datamap: Path) -> Dict[str, int]:
    """
    Returns a dictionary of sheetnames to the number
    of anticipated rows in that sheet according to the
    datamap.
    """
    regex = re.compile(r"(^[a-zA-Z]+)(\d+)$")
    sheet_maxes = {}
    cellrefs_by_sheet = defaultdict(list)
    dmls = datamap_reader(datamap)
    sheet_names = {line.sheet for line in dmls}
    for name in sheet_names:
        for dml in dmls:
            if dml.sheet == name:
                cellrefs_by_sheet[name].append(dml.cellref)
    for s, cellref in cellrefs_by_sheet.items():
        rows = []
        for c in cellref:
            rows.append(int(regex.search(c).groups()[1]))  # type: ignore
        sheet_maxes[s] = max(rows)
    return sheet_maxes
